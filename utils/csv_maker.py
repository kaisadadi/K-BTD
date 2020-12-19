import json
import pickle
import pandas as pd


class Node(object):
    def __init__(self, children, father, ID, name):
        self.children = children
        self.father = father
        self.name = name
        self.ID = ID

    def get_children(self):
        return self.children

    def get_father(self):
        return self.father

    def get_name(self):
        return self.name

    def add_children(self, child):
        self.children.append(child)

    def show(self):
        print(self.ID, self.name, self.children, self.father)

syp_dict_dir = r"../tree/json/syp_dict.json"
tree_dir = r"../tree/json/tree.pkl"
id2pos_dir = r"../tree/json/id2pos.pkl"
all_id2ch_name_dir = r"../tree/json/all_id2ch_name.json"

disease_num2ICD = {"0": "780.6", "1": "401.9", "2": "462", "3": "558.9", "4": "786.2", "5": "465.9", "6": "784.1", "7": "789.0", "8": "530.1", "9": "272.4",
                   "10": "414.0", "11": "250.0", "12": "276.8", "13": "780.4", "14": "599.0", "15": "787.0", "16": "466.0", "17": "463", "18": "486",
                   "19": "440", "20": "475", "21": "784", "22": "434.1", "23": "564.0", "24": "786.5", "25": "882", "26": "493.1", "27": "540.9", "28": "611",
                   "29": "592.9", "30": "724.2", "31": "356.9", "32": "41", "33": "845", "34": "785.1", "35": "460", "36": "873", "37": "276.1", "38": "780.5",
                   "39": "496", "40": "436", "41": "585", "42": "571.9", "43": "351.8", "44": "477.9", "45": "427.3", "46": "787.3", "47": "285.9", "48": "262",
                   "49": "428", "50": "493", "51": "535.0", "52": "530.8", "53": "491", "54": "592.1", "55": "574", "56": "933", "57": "728.8", "58": "536.8",
                   "59": "518.8", "60": "784.7", "61": "9.1", "62": "592.0", "63": "995.3", "64": "35", "65": "560.9", "66": "786.0", "67": "788.0", "68": "600.0",
                   "69": "162.9", "70": "578", "71": "511.9", "72": "575.0", "73": "431", "74": "435.9", "75": "464.0", "76": "577.0", "77": "864", "78": "2",
                   "79": "599.7", "80": "493.2", "81": "412", "82": "351.0", "83": "573", "84": "724.5", "85": "784.0", "86": "535", "87": "536", "88": "269.2",
                   "89": "275.4", "90": "386.1", "91": "274", "92": "723.1", "93": "348.2", "94": "591", "95": "413", "96": "410.9", "97": "276.7", "98": "345.9", "99": "722.1"}   #前100的疾病编码



syp_dict = json.load(open(syp_dict_dir, "r"))   #中文名，symptom
tree = pickle.load(open(tree_dir, "rb"))
id2pos = pickle.load(open(id2pos_dir, "rb"))
all_id2ch_name = json.load(open(all_id2ch_name_dir, "r"))   #icd9 -> 中文名字


#[ICD，英文名，中文名，是否路径上，diagnosis]
csv_list = []
visited = {}   #存放访问过的节点的pos
for num in disease_num2ICD.keys():
    ICD9 = disease_num2ICD[num]
    leaf_pos = id2pos[ICD9]
    father = []
    node = leaf_pos
    while node != None:
        father.append(node)
        node = tree[node].father
    for fa in father:
        if fa not in visited.keys():
            visited[fa] = 0
            csv_list.append({"中文名": syp_dict[tree[fa].ID]['name'], "ICD-9": tree[fa].ID, "key": 1, "diagnosis": syp_dict[tree[fa].ID]["syptom"]})
        for child in tree[fa].children:
            if child not in visited.keys():
                visited[child] = 0
                csv_list.append({"中文名": syp_dict[tree[child].ID]['name'], "ICD-9": tree[child].ID, "key": 1, "diagnosis": syp_dict[tree[child].ID]["syptom"]})

df = pd.DataFrame(csv_list, columns=["中文名", "ICD-9", "key", "diagnosis"])

df.to_excel("../tree/json/human_note.xlsx", index = False)

'''
from openpyxl import *

wb = Workbook()
ws = wb.create_sheet(title="data")

key_match = {0: "中文名", 1: "ICD-9", 2: "key", 3: "diagnosis"}

for idx, item in enumerate(csv_list):

    for k in range(4):
        ws.cell(idx + 1, k + 1)'''


